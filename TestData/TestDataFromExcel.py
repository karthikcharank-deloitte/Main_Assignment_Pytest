import openpyxl


class TestDataFromExcel:

    @staticmethod
    def get_testData(sheet_name):
        book = openpyxl.load_workbook("../TestData/Pytest_TestData.xlsx")
        sheet = book.get_sheet_by_name(sheet_name)
        list_data = []
        for i in range(2, sheet.max_row + 1):
            data_dict = {}
            for j in range(1, sheet.max_column + 1):
                data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            list_data.append(data_dict)
        return list_data
